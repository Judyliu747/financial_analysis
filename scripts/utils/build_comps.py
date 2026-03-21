"""
SMIC Semiconductor Foundry Comparable Company Analysis
Builds an institutional-grade comps Excel model (FY2024 data, market data ~Q1 2026)
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# ─── COLORS ─────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F4E79"
LIGHT_BLUE = "D9E2F3"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"
INPUT_BLUE = "2E74B5"   # blue text = hard-coded input
BLACK      = "000000"
SUBTEXT    = "595959"

# ─── STYLE HELPERS ───────────────────────────────────────────────────────────
def section_hdr(ws, row, col_start, col_end, text):
    """Dark-blue full-width section label."""
    ws.merge_cells(f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}")
    c = ws.cell(row=row, column=col_start, value=text)
    c.font  = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill  = PatternFill(fill_type="solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")

def col_hdr(ws, row, col, text):
    """Light-blue column header."""
    c = ws.cell(row=row, column=col, value=text)
    c.font  = Font(name="Calibri", bold=True, size=10, color=BLACK)
    c.fill  = PatternFill(fill_type="solid", fgColor=LIGHT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def inp(ws, row, col, value, note=None):
    """Blue-text hard-coded input cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font  = Font(name="Calibri", size=10, color=INPUT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if note:
        c.comment = Comment(note, "Comps Model")
    return c

def frm(ws, row, col, formula, fmt=None):
    """Black-text formula cell."""
    c = ws.cell(row=row, column=col, value=formula)
    c.font  = Font(name="Calibri", size=10, color=BLACK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        c.number_format = fmt
    return c

def stat_row(ws, row, col_start, col_end, label):
    """Grey statistics label + fill for the whole row."""
    c = ws.cell(row=row, column=col_start, value=label)
    c.font  = Font(name="Calibri", bold=True, size=10, color=BLACK)
    c.fill  = PatternFill(fill_type="solid", fgColor=LIGHT_GREY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor=LIGHT_GREY)

def stat_frm(ws, row, col, formula, fmt=None):
    """Grey statistics formula cell."""
    c = ws.cell(row=row, column=col, value=formula)
    c.font  = Font(name="Calibri", size=10, color=BLACK)
    c.fill  = PatternFill(fill_type="solid", fgColor=LIGHT_GREY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        c.number_format = fmt
    return c

def nm_cell(ws, row, col, note=None):
    """'NM' (not meaningful) cell."""
    c = ws.cell(row=row, column=col, value="NM")
    c.font  = Font(name="Calibri", size=10, color=SUBTEXT, italic=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if note:
        c.comment = Comment(note, "Comps Model")
    return c

def company_label(ws, row, name):
    c = ws.cell(row=row, column=1, value=name)
    c.font = Font(name="Calibri", bold=True, size=10, color=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center")


# ─── DATA ────────────────────────────────────────────────────────────────────
# Operating data (FY2024, USD millions)
#  name, ticker, revenue, rev_growth, gross_profit, ebitda, net_income, capex
OP_DATA = [
    ("TSMC",            "TSM",      88_340,  0.300,  49_559, 60_532, 36_665, 32_000),
    ("SMIC",            "0981.HK",   8_030,  0.270,   1_445,  4_380,    493,  7_330),
    ("UMC",             "UMC",       7_085, -0.025,   2_049,  2_950,  1_488,  2_200),
    ("GlobalFoundries", "GFS",       6_750, -0.090,   1_654,  2_475,   None,  1_200),
    ("Hua Hong Semi",   "1347.HK",   2_000, -0.123,     204,    500,     58,    900),
    ("Tower Semi",      "TSEM",      1_440,  0.009,     340,    458,    205,    320),
]

# Valuation data (market data ~Q1 2026, USD millions)
#  name, ticker, market_cap, enterprise_value
VAL_DATA = [
    ("TSMC",            "TSM",      850_000, 810_000),
    ("SMIC",            "0981.HK",   86_450,  92_000),
    ("UMC",             "UMC",       26_530,  18_690),
    ("GlobalFoundries", "GFS",       25_920,  24_530),
    ("Hua Hong Semi",   "1347.HK",   23_100,  24_000),
    ("Tower Semi",      "TSEM",      14_040,  12_000),
]

OP_SOURCES = {
    "TSMC":
        "Source: TSMC FY2024 Annual Report (investor.tsmc.com). "
        "Revenue TWD 2,894.31B ÷ 32.8 TWD/USD = $88.3B. "
        "EBITDA $60.5B per MacroTrends/company disclosure (68.5% margin). "
        "Net Income TWD 1,173.27B ÷ 32 = $36.7B.",
    "SMIC":
        "Source: SMIC FY2024 Annual Results (smics.com/en/site/news_read/7919). "
        "Official USD figures. Revenue +27.0% YoY to $8,029.9M. "
        "EBITDA $4,379.7M (+7.8% YoY) per IFRS. CapEx $7.33B.",
    "UMC":
        "Source: UMC Q4 2024 Earnings Release (businesswire.com, Jan 21, 2025). "
        "Revenue NT$226.7B ÷ 32 = ~$7.09B. Gross margin 28.9%. "
        "EBITDA estimated: operating income (~18.5% margin) + D&A (~$1.6B est.).",
    "GlobalFoundries":
        "Source: GlobalFoundries FY2024 Financial Results (investors.gf.com, Feb 11, 2025). "
        "Revenue $6,750M (–9% YoY). Gross margin 24.5%. "
        "Non-IFRS Adjusted EBITDA $2,475M. GAAP net loss $262M due to $935M asset impairment.",
    "Hua Hong Semi":
        "Source: Hua Hong Semiconductor FY2024 results (thestandard.com.hk). "
        "Revenue –12.3% YoY to ~$2,000M. Gross margin 10.2% (–11.1pp YoY). "
        "EBITDA estimated ~$500M (25% margin); heavy D&A from recent capacity expansion.",
    "Tower Semi":
        "Source: Tower Semiconductor FY2024 Results (globenewswire.com, Feb 10, 2025). "
        "Revenue $1,440M (+0.9% YoY). Gross margin 23.6%. "
        "EBITDA $457.6M (31.8% margin) per company disclosure.",
}

VAL_SOURCES = {
    "TSMC":
        "Market Cap ~$850B: TSMC ~5.18B shares × ~$164/ADS (end-2024 est.). "
        "EV = Mkt Cap – net cash ~$40B. Note: price has risen further in early 2026.",
    "SMIC":
        "Market Cap ~$86.5B: HKG:0981 at HKD 63.75/share × 10.64B shares ÷ 7.83. "
        "EV = Mkt Cap + estimated net debt $5.55B (total debt ~$15B, cash ~$9.45B). "
        "Premium vs. global peers reflects SMIC's strategic importance to China's chip self-sufficiency.",
    "UMC":
        "Source: StockAnalysis.com / Yahoo Finance (Jan 2026). "
        "Market Cap $26.53B, Enterprise Value $18.69B (implies ~$7.8B net cash).",
    "GlobalFoundries":
        "Source: Yahoo Finance key statistics. Market Cap $25.92B, EV $24.53B. "
        "EV/EBITDA based on Non-IFRS Adj. EBITDA $2,475M.",
    "Hua Hong Semi":
        "Market Cap ~$23.1B: HKG:1347 at 181.08B HKD ÷ 7.83. "
        "EV est. ~$24B (net debt ~$900M). Extreme P/E reflects China strategic-asset premium.",
    "Tower Semi":
        "Market Cap $14.04B per companiesmarketcap.com (March 2026). "
        "EV ~$12B estimated (net cash ~$2B from strong FCF generation).",
}

GFS_NI_NOTE = (
    "GFS reported GAAP net loss of $262M in FY2024, driven by a $935M "
    "impairment charge on long-lived assets at its Malta, NY facility. "
    "Non-GAAP net income was $870M. P/E is NM on a GAAP basis."
)


# ─── BUILD WORKBOOK ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Comps Analysis"

NCOLS = 10   # A–J
NVCOLS = 7   # A–G (valuation section)

# ── TITLE BLOCK (rows 1–3) ────────────────────────────────────────────────────
ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
ws["A1"] = "SEMICONDUCTOR FOUNDRY — COMPARABLE COMPANY ANALYSIS"
ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
ws["A1"].fill      = PatternFill(fill_type="solid", fgColor=DARK_BLUE)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
ws["A2"] = ("TSMC (TSM)  •  SMIC (0981.HK)  •  UMC (UMC)  •  "
            "GlobalFoundries (GFS)  •  Hua Hong Semiconductor (1347.HK)  •  Tower Semiconductor (TSEM)")
ws["A2"].font      = Font(name="Calibri", size=10, color=BLACK)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells(f"A3:{get_column_letter(NCOLS)}3")
ws["A3"] = "As of Q1 2026  |  FY2024 Financial Data  |  All figures in USD Millions except margins and ratios"
ws["A3"].font      = Font(name="Calibri", italic=True, size=10, color=SUBTEXT)
ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

# ── SECTION 1: OPERATING STATISTICS (rows 5–19) ───────────────────────────────
# Row 4: spacer (empty)
# Row 5: section header
section_hdr(ws, 5, 1, NCOLS, "OPERATING STATISTICS & FINANCIAL METRICS")

# Row 6: column headers
OP_HDRS = [
    "Company", "Ticker",
    "Revenue\n(FY2024, $M)", "Revenue\nGrowth (YoY)",
    "Gross Profit\n($M)", "Gross\nMargin",
    "EBITDA\n($M)", "EBITDA\nMargin",
    "Net Income\n($M)", "CapEx\n($M)",
]
for ci, hdr in enumerate(OP_HDRS, 1):
    col_hdr(ws, 6, ci, hdr)

# Rows 7–12: company data
OP_ROW_START = 7
for ri, (name, ticker, rev, rev_g, gp, ebitda, ni, capex) in enumerate(OP_DATA, OP_ROW_START):
    src = OP_SOURCES.get(name, "")
    company_label(ws, ri, name)
    inp(ws, ri, 2, ticker)
    inp(ws, ri, 3, rev,      src).number_format = "#,##0"
    inp(ws, ri, 4, rev_g,   src).number_format = "0.0%"
    inp(ws, ri, 5, gp,       src).number_format = "#,##0"

    # Gross Margin = GP / Revenue (formula)
    frm(ws, ri, 6, f"=E{ri}/C{ri}", "0.0%")

    inp(ws, ri, 7, ebitda, src).number_format = "#,##0"

    # EBITDA Margin = EBITDA / Revenue (formula)
    frm(ws, ri, 8, f"=G{ri}/C{ri}", "0.0%")

    # Net Income – NM for GlobalFoundries
    if name == "GlobalFoundries":
        nm_cell(ws, ri, 9, GFS_NI_NOTE)
    else:
        inp(ws, ri, 9, ni, src).number_format = "#,##0"

    inp(ws, ri, 10, capex, src).number_format = "#,##0"

OP_ROW_END = OP_ROW_START + len(OP_DATA) - 1  # = 12

# Row 13: blank separator

# Rows 14–18: statistics
STAT_DEFS = [
    ("Maximum",         "MAX",      None),
    ("75th Percentile", "QUARTILE", 3),
    ("Median",          "MEDIAN",   None),
    ("25th Percentile", "QUARTILE", 1),
    ("Minimum",         "MIN",      None),
]

def make_stat_formula(func, quart, col_letter, r1, r2):
    rng = f"{col_letter}{r1}:{col_letter}{r2}"
    if func == "MAX":      return f"=MAX({rng})"
    if func == "MIN":      return f"=MIN({rng})"
    if func == "MEDIAN":   return f"=MEDIAN({rng})"
    if func == "QUARTILE": return f"=QUARTILE({rng},{quart})"

STAT_OP_COLS = {4: "D", 6: "F", 8: "H"}   # Rev Growth, Gross Margin, EBITDA Margin

STAT_ROW_START = 14
for si, (label, func, quart) in enumerate(STAT_DEFS, STAT_ROW_START):
    stat_row(ws, si, 1, NCOLS, label)
    for col_idx, col_ltr in STAT_OP_COLS.items():
        f = make_stat_formula(func, quart, col_ltr, OP_ROW_START, OP_ROW_END)
        stat_frm(ws, si, col_idx, f, "0.0%")

STAT_ROW_END = STAT_ROW_START + len(STAT_DEFS) - 1  # = 18

# ── SECTION 2: VALUATION MULTIPLES (rows 21–35) ───────────────────────────────
# Row 20: spacer

VAL_SEC_ROW = 21
section_hdr(ws, VAL_SEC_ROW, 1, NVCOLS, "VALUATION MULTIPLES")

# Row 22: column headers
VAL_HDRS = [
    "Company", "Ticker",
    "Market Cap\n($M)", "Enterprise\nValue ($M)",
    "EV / Revenue", "EV / EBITDA", "P / E",
]
for ci, hdr in enumerate(VAL_HDRS, 1):
    col_hdr(ws, VAL_SEC_ROW + 1, ci, hdr)

# Rows 23–28: valuation data
VAL_ROW_START = 23
OP_OFFSET = VAL_ROW_START - OP_ROW_START  # = 16; val_row - offset = op_row

for ri, (name, ticker, mktcap, ev) in enumerate(VAL_DATA, VAL_ROW_START):
    op_row = ri - OP_OFFSET
    src    = VAL_SOURCES.get(name, "")
    company_label(ws, ri, name)
    inp(ws, ri, 2, ticker)
    inp(ws, ri, 3, mktcap, src).number_format = "#,##0"
    inp(ws, ri, 4, ev,     src).number_format = "#,##0"

    # EV / Revenue  = D{ri} / C{op_row}  (EV ÷ op-section revenue)
    frm(ws, ri, 5, f"=D{ri}/C{op_row}", '0.0"x"')

    # EV / EBITDA = D{ri} / G{op_row}
    frm(ws, ri, 6, f"=D{ri}/G{op_row}", '0.0"x"')

    # P/E = Market Cap / Net Income  (NM for GFS whose I{op_row} = text "NM")
    if name == "GlobalFoundries":
        nm_cell(ws, ri, 7,
                "P/E NM: GFS GAAP net loss $262M in FY2024 due to $935M asset impairment.")
    else:
        frm(ws, ri, 7,
            f'=IFERROR(C{ri}/I{op_row},"NM")', '0.0"x"')

VAL_ROW_END = VAL_ROW_START + len(VAL_DATA) - 1  # = 28

# Row 29: blank separator

# Rows 30–34: valuation statistics
STAT_VAL_COLS = {5: "E", 6: "F", 7: "G"}   # EV/Rev, EV/EBITDA, P/E

STAT_VAL_ROW_START = 30
for si, (label, func, quart) in enumerate(STAT_DEFS, STAT_VAL_ROW_START):
    stat_row(ws, si, 1, NVCOLS, label)
    for col_idx, col_ltr in STAT_VAL_COLS.items():
        # For P/E (col G), GFS cell is text "NM" — Excel MAX/MIN/MEDIAN ignore text automatically
        f = make_stat_formula(func, quart, col_ltr, VAL_ROW_START, VAL_ROW_END)
        stat_frm(ws, si, col_idx, f, '0.0"x"')

# ── SECTION 3: NOTES (rows 37+) ───────────────────────────────────────────────
NOTE_SEC_ROW = 37
ws.merge_cells(f"A{NOTE_SEC_ROW}:{get_column_letter(NCOLS)}{NOTE_SEC_ROW}")
ws[f"A{NOTE_SEC_ROW}"] = "NOTES & METHODOLOGY"
ws[f"A{NOTE_SEC_ROW}"].font      = Font(name="Calibri", bold=True, size=11, color=WHITE)
ws[f"A{NOTE_SEC_ROW}"].fill      = PatternFill(fill_type="solid", fgColor=DARK_BLUE)
ws[f"A{NOTE_SEC_ROW}"].alignment = Alignment(horizontal="left", vertical="center")

NOTES = [
    ("Data Sources",
     "FY2024 financials sourced from official earnings releases and annual reports: "
     "TSMC (investor.tsmc.com), SMIC (smics.com), UMC (umc.com/ir), "
     "GlobalFoundries (investors.gf.com), Hua Hong (huahonggrace.com), "
     "Tower Semiconductor (globenewswire.com). Market cap / EV data as of ~Q1 2026."),
    ("Currency",
     "All figures in USD Millions. TWD figures (TSMC, UMC) converted at ~32 TWD/USD. "
     "HKD figures (SMIC, Hua Hong) converted at 7.83 HKD/USD. CNY at 7.18 CNY/USD."),
    ("EBITDA",
     "SMIC and GFS EBITDA from official disclosures. "
     "TSMC EBITDA $60.5B per MacroTrends (68.5% margin, consistent with high D&A on ~$300B+ PP&E). "
     "UMC and Hua Hong EBITDA estimated as Operating Income + D&A. "
     "Note: foundry EBITDA margins appear high (~40–70%) because D&A on capital-intensive PP&E is added back."),
    ("Enterprise Value",
     "EV = Market Cap + Net Debt (Total Debt – Cash & Equivalents). "
     "TSMC and UMC carry significant net cash positions, reducing EV below market cap. "
     "SMIC and Hua Hong carry net debt, reflecting ongoing capacity-expansion capex programs."),
    ("China Strategic Premium",
     "SMIC and Hua Hong trade at substantial premiums to global peers on P/E (175x and ~400x respectively). "
     "This reflects China's national push for semiconductor self-sufficiency, government support expectations, "
     "and scarcity of domestically listed advanced foundry assets amid US export controls — "
     "not a reflection of near-term earnings power."),
    ("GlobalFoundries P/E (NM)",
     "GFS reported a GAAP net loss of $262M in FY2024 due to a $935M impairment charge "
     "on long-lived assets at its Malta, NY facility. Non-GAAP net income was $870M. "
     "EV/EBITDA uses Non-IFRS Adjusted EBITDA of $2,475M per GFS disclosure."),
    ("CapEx Intensity",
     "Semiconductor foundry is highly capital-intensive. SMIC FY2024 CapEx of $7.33B "
     "represents 91% of revenue — among the most aggressive expansions globally. "
     "TSMC CapEx ~$32B (36% of revenue). CapEx/Revenue ratio drives future D&A and should be "
     "considered alongside EBITDA when assessing sustainable free cash flow."),
    ("Peer Selection",
     "Peer group limited to pure-play contract IC manufacturers. "
     "Samsung Foundry excluded (division of diversified conglomerate). "
     "Intel Foundry excluded (pre-scale, transition phase with insufficient comparable history). "
     "Powerchip (non-public) and PSMC excluded (limited disclosure)."),
]

for ni, (title, content) in enumerate(NOTES, NOTE_SEC_ROW + 1):
    c_title = ws.cell(row=ni, column=1, value=title)
    c_title.font      = Font(name="Calibri", bold=True, size=10, color=BLACK)
    c_title.fill      = PatternFill(fill_type="solid", fgColor=LIGHT_GREY)
    c_title.alignment = Alignment(horizontal="left", vertical="top")

    ws.merge_cells(f"B{ni}:{get_column_letter(NCOLS)}{ni}")
    c_body = ws.cell(row=ni, column=2, value=content)
    c_body.font      = Font(name="Calibri", size=10, color=BLACK)
    c_body.fill      = PatternFill(fill_type="solid", fgColor=LIGHT_GREY)
    c_body.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[ni].height = 48

# ─── DIMENSIONS ──────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 32
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 18
ws.row_dimensions[5].height = 22
ws.row_dimensions[6].height = 38
ws.row_dimensions[VAL_SEC_ROW].height = 22
ws.row_dimensions[VAL_SEC_ROW + 1].height = 38
ws.row_dimensions[NOTE_SEC_ROW].height = 22

for r in list(range(OP_ROW_START, OP_ROW_END + 1)) + \
         list(range(STAT_ROW_START, STAT_ROW_END + 1)) + \
         list(range(VAL_ROW_START, VAL_ROW_END + 1)) + \
         list(range(STAT_VAL_ROW_START, STAT_VAL_ROW_START + len(STAT_DEFS))):
    ws.row_dimensions[r].height = 20

# Column widths
COL_WIDTHS = {1: 20, 2: 11, 3: 16, 4: 14, 5: 15, 6: 12, 7: 14, 8: 13, 9: 14, 10: 12}
for col, w in COL_WIDTHS.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "C7"

# ─── SAVE ─────────────────────────────────────────────────────────────────────
out_dir  = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(out_dir, "SMIC_Foundry_Comps_FY2024.xlsx")
wb.save(out_path)
print(f"Saved → {out_path}")
